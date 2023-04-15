from tkinter import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

root = Tk()
root.title('OpenpyXL')
root.geometry('500x900')


wb = Workbook()
wb = load_workbook('Table.xlsx')
ws = wb.active

column_a = ws['A']
column_b = ws['B']

print(column_a)


root.mainloop()